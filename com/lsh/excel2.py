# -*- coding: utf-8 -*-
# @Time : 2020/4/23 16:38
# @Author : LSH
# @Desc : 
# @Version : 1.0.0

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('学生明细表.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
